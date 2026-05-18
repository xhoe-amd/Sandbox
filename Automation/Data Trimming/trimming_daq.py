import numpy as np
import statistics
import csv
import matplotlib.pyplot as plt
from openpyxl import load_workbook, Workbook
from openpyxl.chart import LineChart, Reference

# def roc_limits_mad(values, window=5, threshold=3):
#     rates = [
#         values[i + window - 1] - values[i]
#         for i in range(len(values) - window + 1)
#     ]

#     r = np.array(rates, dtype=float)
#     med = np.median(r)
#     mad = np.median(np.abs(r - med))

#     upper = med + threshold * 1.4826 * mad
#     lower = med - threshold * 1.4826 * mad

#     return lower, upper, r

# lower, upper, rates = roc_limits_mad(val, window=5)

# median_roc = statistics.median(r[1] for r in roc_list)
# mad = statistics.median(abs(x - median_roc) for x in (r[1] for r in roc_list))

# n_threshold = median_roc - 3 * mad

# def trend_minus1_to_1(values):
#     s = np.sort(values)
#     n = len(s)

#     if n < 2 or s[-1] == s[0]:
#         return 0.0   # no trend possible,

#     x = np.arange(1, n + 1)
#     x_mean = (n + 1) / 2
#     s_mean = s.mean()

#     numerator = np.sum((x - x_mean) * (s - s_mean))
#     normalization = (n * (n**2 - 1)) / 12

#     return (numerator / normalization) / (s[-1] - s[0])

csv_file = "0_daq_ddc96a15-5f79-4224-8e29-f9d4765f2fae.power.csv"
target_data = "APU_Total_P=APU_VDDCR_PCCX_P+APU_VDDCR_ECCX02_P+APU_VDDCR_ECCX13_P+APU_VDDCR_AIE_P+APU_VDDCR_GFX_P+APU_VDDCR_SOC_P+ROC_DRAM_Total_P"


column_index = None
start_row = None
rows = []

# Read CSV into memory
with open(csv_file, newline='', encoding='utf-8') as f:
    reader = list(csv.reader(f))
    rows = reader

# Find the cell containing target_data
for row_idx, row in enumerate(rows):
    for col_idx, cell in enumerate(row):
        if cell == target_data:
            column_index = col_idx
            start_row = row_idx + 2  # start BELOW the cell
            break
    if column_index is not None:
        break

# Pull values below that cell in the same column
data = []

if column_index is not None:
    for row_index, row in enumerate(rows[start_row:], start=start_row):
        if column_index < len(row) and row[column_index] != "":
            data.append((row_index, float(row[column_index])))


def rate_of_change(y, t=None):
    y = np.asarray(y, dtype=float)
    if t is None:
        t = np.arange(len(y), dtype=float)
    else:
        t = np.asarray(t, dtype=float)

    if len(y) < 2:
        return 0.0

    return np.polyfit(t, y, 1)[0]

min_sample_size = 5
max_sample_size = 25
sample_size = max(min_sample_size, min(int(len(data) ** 0.5), max_sample_size)) # min_sample_size
min_val = min(data, key=lambda x: x[1])[1]

count = len(data) - sample_size

roc_list = []
trend_list = []

for v in range(count):
    sample = []

    for i in range(sample_size):
        sample.append(data[v + i][1])

    roc = rate_of_change(sample)
    roc_list.append((sample, roc))

mean_roc = statistics.mean(r[1] for r in roc_list)
std_roc  = statistics.stdev(r[1] for r in roc_list)

p_threshold = mean_roc + 2 * std_roc
n_threshold = mean_roc - 2 * std_roc

## TODO Handle very sudden spike in small sample (e.g. from 3v -> 15v -> 3v)

print(p_threshold)
print(n_threshold)
print(min_val)

result = []
detect = []

start_sample_check = [min_val] + roc_list[0][0]

if rate_of_change(start_sample_check) > p_threshold:
    detect.append(0)

for r in range(len(roc_list)):
    if len(detect) >= 2:
        result.append(list(detect))
        detect.clear()

    # Conditions:
    # 1. Check if starting increase spike is found, if yes any increase spike after this will be ignored until decrease spike is found
    # 2. Check if sample ROC is larger than increase ROC threshold (General detection)
    # 3. Check ROC of last 2 values of sample is actually larger than increase ROC threshold (Specific detection)
    # 4. If all previous conditions are satisfied, check for ROC of all previous values from current index to make sure it's stable (smaller than increase ROC threshold)
    if len(detect) == 0 and roc_list[r][1] > p_threshold and (rate_of_change(roc_list[r][0][-abs(min_sample_size):])) > p_threshold:
        if rate_of_change([v for _, v in data[:r]]) < p_threshold:
            print(roc_list[r][0])
            print(rate_of_change(roc_list[r][0][-abs(min_sample_size):]))
            detect.append(data[r + sample_size][0])

    # Conditions:
    # 1. Check if starting increase spike is already found, if yes then the first new decrease spike detected will be considered, if increase spike is not yet found, ignore any decrease spike
    # 2. Check if sample ROC is smaller than decrease ROC threshold (General detection)
    # 3. Check ROC of last 2 values of sample is actually smaller than decrease ROC threshold (Specific detection)
    # 4. Check ROC of last value of current sample and next sample remains stable (between increase ROC and decrease ROC)
    if len(detect) == 1 and roc_list[r][1] < n_threshold and (rate_of_change(roc_list[r][0][-abs(min_sample_size):])) < n_threshold and (((rate_of_change([roc[0][-1] for roc in roc_list[r:r+2]])) < p_threshold and (rate_of_change([roc[0][-1] for roc in roc_list[r:r+2]])) > n_threshold) if r < len(roc_list) else True):
        print(roc_list[r][0])
        print(rate_of_change(roc_list[r][0][-abs(min_sample_size):]))
        # Deducting 3 because condition checking will also ensure last 2 values ROC are less than increase ROC (meaning the values are now stable, which should represents the workloads are not running anymore), and 1 more due to excel rows index is 0 based
        detect.append(data[r + sample_size - 3][0])

if len(detect) == 1:
    detect.append(len(data))
    result.append(detect)
elif len(result) == 0:
    result.append([0, len(data)])

print(result)

# for r in result:
#     x = []  # row indices
#     y = []  # values

#     with open(csv_file, newline='', encoding='utf-8') as f:
#         rows = list(csv.reader(f))

#     print(r)
#     for row_index in range(r[0], r[1] + 1):
#         row = rows[row_index]
#         if column_index < len(row) and row[column_index] != "":
#             x.append(row_index)
#             y.append(float(row[column_index]))

#     # Plot
#     plt.plot(x, y, marker='o')
#     plt.xlabel("Row Index")
#     plt.ylabel("Value")
#     plt.title("Line Graph for Selected Rows")
#     plt.grid(True)
#     plt.show()



# def infer_type(value):
#     if value == "":
#         return None
#     try:
#         if "." in value:
#             return float(value)
#         return int(value)
#     except ValueError:
#         return value  # keep as text


# output_file = "output_filtered_chart.xlsx"


# xlsx_file = "input.xlsx"


# wb = Workbook()
# ws = wb.active

# with open(csv_file, newline="", encoding="utf-8") as f:
#     reader = csv.reader(f)
#     for row in reader:
#         ws.append([infer_type(cell) for cell in row])

# wb.save(xlsx_file)

# print("CSV converted to XLSX with inferred data types")



# wb = load_workbook(xlsx_file)
# ws = wb.active

# # Convert to Excel indexing
# excel_col = column_index + 1
# excel_start = result[0][0] + 1
# excel_end = result[0][1] + 1

# # ---------------------------
# # Apply AutoFilter to the sheet
# # ---------------------------
# ws.auto_filter.ref = ws.dimensions

# # Filter rows visually (hides rows outside range)
# for row in range(2, ws.max_row + 1):
#     if not (excel_start <= row <= excel_end):
#         ws.row_dimensions[row].hidden = True

# # ---------------------------
# # Create the line chart
# # ---------------------------
# chart = LineChart()
# chart.title = "Filtered Line Graph"
# chart.y_axis.title = "Value"
# chart.x_axis.title = "Row"

# data = Reference(
#     ws,
#     min_col=excel_col,
#     min_row=excel_start,
#     max_row=excel_end
# )

# cats = Reference(
#     ws,
#     min_col=1,
#     min_row=excel_start,
#     max_row=excel_end
# )

# chart.add_data(data, titles_from_data=False)
# chart.set_categories(cats)

# ws.add_chart(chart, "E2")

# # ---------------------------
# # Save file
# # ---------------------------
# wb.save(output_file)

# print("Excel saved with filtered rows and chart.")


# print(trend_minus1_to_1(data))